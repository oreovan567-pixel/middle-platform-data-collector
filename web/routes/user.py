"""用户管理 API"""
import io
from collections import OrderedDict

from flask import Blueprint, request, jsonify, session, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from models.user import User
from models.school import School

user_bp = Blueprint("users", __name__)


def _require_admin():
    if not session.get("is_admin") and session.get("role") != "super_admin":
        return jsonify({"error": "需要超级管理员权限"}), 403
    return None


@user_bp.route("/", methods=["GET"])
def list_users():
    users = User.get_all()
    return jsonify({"users": [u.to_dict(include_passwords=True) for u in users]})


@user_bp.route("/me", methods=["GET"])
def get_me():
    user_id = session.get("user_id")
    if not user_id:
        return jsonify({"error": "未登录"}), 401
    user = User.get_by_id(user_id)
    if not user:
        return jsonify({"error": "用户不存在"}), 404
    return jsonify({"user": user.to_dict(include_passwords=True)})


@user_bp.route("/me", methods=["PUT"])
def update_me():
    user_id = session.get("user_id")
    if not user_id:
        return jsonify({"error": "未登录"}), 401
    user = User.get_by_id(user_id)
    if not user:
        return jsonify({"error": "用户不存在"}), 404

    data = request.get_json()
    if not data:
        return jsonify({"error": "请求体不能为空"}), 400

    # 普通用户只能修改自己的凭证
    if data.get("lida_username") is not None:
        user.lida_username = data["lida_username"]
    if data.get("lida_password") is not None:
        user.lida_password = data["lida_password"]
    if data.get("grafana_username") is not None:
        user.grafana_username = data["grafana_username"]
    if data.get("grafana_password") is not None:
        user.grafana_password = data["grafana_password"]
    if data.get("main_site_username") is not None:
        user.main_site_username = data["main_site_username"]
    if data.get("main_site_password") is not None:
        user.main_site_password = data["main_site_password"]
    if data.get("password"):
        user.password = data["password"]

    user.save()
    return jsonify({"message": "更新成功", "user": user.to_dict()})


@user_bp.route("/", methods=["POST"])
def create_user():
    err = _require_admin()
    if err:
        return err

    data = request.get_json()
    if not data or not data.get("username"):
        return jsonify({"error": "用户名为必填"}), 400

    existing = User.get_by_username(data["username"])
    if existing:
        uname = data["username"]
        return jsonify({"error": f"用户名 '{uname}' 已存在"}), 409

    user = User(
        username=data["username"],
        password=data.get("password", ""),
        lida_username=data.get("lida_username", ""),
        lida_password=data.get("lida_password", ""),
        grafana_username=data.get("grafana_username", ""),
        grafana_password=data.get("grafana_password", ""),
        main_site_username=data.get("main_site_username", ""),
        main_site_password=data.get("main_site_password", ""),
        assigned_schools=data.get("assigned_schools", ""),
        is_admin=data.get("is_admin", False),
    )
    user.save()
    return jsonify({"message": "用户创建成功", "user": user.to_dict()}), 201


@user_bp.route("/<int:user_id>", methods=["PUT"])
def update_user(user_id):
    user = User.get_by_id(user_id)
    if not user:
        return jsonify({"error": "用户不存在"}), 404

    # 管理员可以改任何人，普通用户只能改自己
    if not session.get("is_admin") and session.get("user_id") != user_id:
        return jsonify({"error": "权限不足"}), 403

    data = request.get_json()
    if not data:
        return jsonify({"error": "请求体不能为空"}), 400

    # 管理员可以改所有字段
    if session.get("is_admin"):
        if data.get("username"):
            user.username = data["username"]
        if data.get("password"):
            user.password = data["password"]
        if data.get("assigned_schools") is not None:
            user.assigned_schools = data["assigned_schools"]
        if data.get("is_admin") is not None:
            user.is_admin = data["is_admin"]

    # 凭证字段所有人都可以改
    for field in ("lida_username", "lida_password", "grafana_username",
                  "grafana_password", "main_site_username", "main_site_password"):
        if data.get(field) is not None:
            setattr(user, field, data[field])

    user.save()
    return jsonify({"message": "更新成功", "user": user.to_dict()})


# ===== 批量导入 =====
_IMPORT_COLUMNS = [
    ("用户名",      "username"),
    ("学校名称",    "school_name"),
    ("Metabase学校ID", "metabase_school_id"),
    ("Grafana名称", "grafana_name"),
    ("主站名称",    "main_site_name"),
    ("学段",        "xueduan"),
    ("年级",        "nianji"),
]

_HEADER_NAMES = [col[0] for col in _IMPORT_COLUMNS]


@user_bp.route("/import-template", methods=["GET"])
def download_import_template():
    """下载导入模板 Excel"""
    err = _require_admin()
    if err:
        return err

    wb = Workbook()
    ws = wb.active
    ws.title = "用户及学校导入"

    # 表头样式
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0891B2")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, name in enumerate(_HEADER_NAMES, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    widths = [14, 22, 22, 22, 22, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    examples = [
        ["张三", "宜宾市第一小学", "宜宾一小", "yibin_no1", "ybsdyxx", "小学", "三年级"],
        ["张三", "宜宾市第二小学", "宜宾二小", "yibin_no2", "ybsdexx", "小学", "四年级"],
        ["李四", "成都实验中学",     "成都实验", "cd_syzx",   "cdsyzz",  "初中", "初一"],
    ]
    data_font = Font(name="微软雅黑", size=10)
    for row_idx, row_data in enumerate(examples, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    note_row = len(examples) + 3
    notes = [
        "【填写说明】",
        "1. 用户名：登录系统使用的账号名（同一用户多所学校时，用户名重复填写多行即可）",
        "2. 学校名称：在系统中显示的学校名称（必填，不可重复）",
        "3. Lida名称：该平台中对应的学校名称",
        "4. Grafana名称：Grafana 系统中对应的学校名称",
        "5. 主站名称：主站系统中对应的学校名称",
        "6. 学段：如 小学、初中、高中",
        "7. 年级：如 三年级、初一、高一",
        "",
        "注意：示例行请删除后再上传。",
    ]
    note_font = Font(name="微软雅黑", size=10, color="666666")
    for i, line in enumerate(notes):
        cell = ws.cell(row=note_row + i, column=1, value=line)
        cell.font = note_font
        ws.merge_cells(start_row=note_row + i, start_column=1,
                       end_row=note_row + i, end_column=len(_HEADER_NAMES))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="用户及学校导入模板.xlsx",
    )


@user_bp.route("/import", methods=["POST"])
def batch_import():
    """批量导入用户及学校信息"""
    err = _require_admin()
    if err:
        return err

    f = request.files.get("file")
    if not f:
        return jsonify({"error": "请上传 Excel 文件"}), 400

    try:
        wb = load_workbook(f, read_only=True, data_only=True)
    except Exception:
        return jsonify({"error": "文件格式错误，请上传 .xlsx 文件"}), 400

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        return jsonify({"error": "文件中没有数据行"}), 400

    # 解析每行
    parsed_rows = []
    errors = []
    for i, row in enumerate(rows, start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        vals = [str(c).strip() if c is not None else "" for c in row[:len(_IMPORT_COLUMNS)]]
        username = vals[0]
        school_name = vals[1]
        if not username:
            errors.append(f"第{i}行：用户名不能为空")
            continue
        if not school_name:
            errors.append(f"第{i}行：学校名称不能为空")
            continue
        parsed_rows.append({
            "username": username,
            "school_name": school_name,
            "metabase_school_id": vals[2],
            "grafana_name": vals[3],
            "main_site_name": vals[4],
            "xueduan": vals[5],
            "nianji": vals[6],
        })

    if not parsed_rows:
        return jsonify({"error": "没有有效数据行", "details": errors}), 400

    # 按用户名分组（保持顺序）
    user_schools = OrderedDict()
    for r in parsed_rows:
        user_schools.setdefault(r["username"], []).append(r)

    created_users = []
    existing_users = []
    created_schools = []
    updated_schools = []
    current_user_id = session.get("user_id")

    for username, schools in user_schools.items():
        user = User.get_by_username(username)
        if user:
            existing_users.append(username)
        else:
            user = User(username=username, password="", is_admin=False)
            user.save()
            created_users.append(username)

        school_names = [s["school_name"] for s in schools]
        user.assigned_schools = ",".join(school_names)
        user.save()

        for s in schools:
            existing_school = School.get_by_name(s["school_name"])
            school = School(
                name=s["school_name"],
                lida_name=s["school_name"],
                metabase_school_id=s["metabase_school_id"],
                grafana_name=s["grafana_name"],
                main_site_name=s["main_site_name"],
                xueduan=s["xueduan"],
                nianji=s["nianji"],
                owner_id=current_user_id,
            )
            if existing_school:
                school.id = existing_school.id
                school.save()
                updated_schools.append(s["school_name"])
            else:
                school.save()
                created_schools.append(s["school_name"])

    wb.close()

    result = {
        "message": "导入完成",
        "summary": {
            "total_rows": len(parsed_rows),
            "new_users": len(created_users),
            "existing_users": len(existing_users),
            "new_schools": len(created_schools),
            "updated_schools": len(updated_schools),
        },
        "details": {
            "created_users": created_users,
            "existing_users": existing_users,
            "created_schools": created_schools,
            "updated_schools": updated_schools,
        },
    }
    if errors:
        result["warnings"] = errors
    return jsonify(result)


@user_bp.route("/<int:user_id>", methods=["DELETE"])
def delete_user(user_id):
    err = _require_admin()
    if err:
        return err

    user = User.get_by_id(user_id)
    if not user:
        return jsonify({"error": "用户不存在"}), 404
    if user.username == "admin":
        return jsonify({"error": "不能删除默认管理员"}), 400

    user.delete()
    return jsonify({"message": "用户已删除"})

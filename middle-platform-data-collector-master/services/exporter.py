"""Excel 导出服务"""
import os
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from models.weekly_record import WeeklyRecord
from models.monthly_record import MonthlyRecord


_EXPORT_DIR = Path(__file__).parent.parent / "data" / "exports"

# 表头定义
HEADERS = [
    "学校名称",
    "周次",
    "整体使用率",
    "整体集备",
    "级部集备",
    "学部集备",
    "作业次数",
    "本周活跃教师",
    "本周使用总教师",
    "本周整体活跃度",
    "周活教师比例",
    "日期",
]

# 样式
_HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_CELL_FONT = Font(name="微软雅黑", size=10)
_CELL_ALIGN = Alignment(horizontal="center", vertical="center")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)



def _build_export_name(year, period):
    """
    Build export filename: year年 + period + 数据统计
    """
    parts = [str(year) + "年"]
    if period:
        parts.append(period)
    parts.append("数据统计")
    return re.sub(r'[\\/:*?"<>|\s]', '_', "".join(parts))


def _safe_filename(label: str) -> str:
    """将文本标签转为安全的文件名片段"""
    return re.sub(r'[\\/:*?"<>|\s]', '_', label)


def export_weekly(
    records: list[WeeklyRecord],
    year: int,
    week_number: str,
) -> str:
    """
    将周表记录导出为 Excel 文件。

    Returns:
        导出文件的路径
    """
    _EXPORT_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}年{week_number or '全部'}"[:31]

    # 写入标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"{year}年{week_number or '全部'}数据统计"
    title_cell.font = Font(name="微软雅黑", size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # 写入表头
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER

    # 写入数据
    for row_idx, record in enumerate(records, 3):
        _date = ''
        if hasattr(record, 'collected_at') and record.collected_at:
            _dt = record.collected_at
            if 'T' in _dt:
                _date = _dt.replace('T', ' ')[:16]  # YYYY-MM-DD HH:MM
            else:
                _date = _dt[:16]
        values = [
            record.school_name,
            record.week_number,
            record.overall_usage_rate,
            record.overall_jibei,
            record.grade_jibei,
            record.department_jibei,
            record.homework_count,
            record.weekly_active_teachers,
            record.weekly_total_teachers,
            record.weekly_overall_activity,
            record.weekly_active_ratio,
            _date,
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = _CELL_FONT
            cell.alignment = _CELL_ALIGN
            cell.border = _THIN_BORDER

    # 调整列宽
    col_widths = [18, 10, 12, 12, 12, 12, 12, 14, 14, 14, 12, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # 保存文件
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_label = _safe_filename(week_number)
    filename = _build_export_name(year, week_number) + f"_{timestamp}.xlsx"
    filepath = _EXPORT_DIR / filename
    wb.save(str(filepath))

    return str(filepath)


def export_weekly_range(
    records: list[WeeklyRecord],
    year: int,
    week_labels: list[str],
) -> str:
    """按周次列表导出，每周一个sheet"""
    _EXPORT_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)  # 移除默认sheet

    # 按周次分组
    grouped: dict[str, list[WeeklyRecord]] = {}
    for r in records:
        grouped.setdefault(r.week_number, []).append(r)

    for week in sorted(grouped.keys()):
        week_records = grouped[week]
        ws = wb.create_sheet(title=week[:31])

        # 表头
        for col_idx, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGN
            cell.border = _THIN_BORDER

        # 数据
        for row_idx, record in enumerate(week_records, 2):
            _date = ''
            if hasattr(record, 'collected_at') and record.collected_at:
                _dt = record.collected_at
                if 'T' in _dt:
                    _date = _dt.replace('T', ' ')[:16]  # YYYY-MM-DD HH:MM
                else:
                    _date = _dt[:16]
            values = [
                record.school_name,
                record.week_number,
                record.overall_usage_rate,
                record.overall_jibei,
                record.grade_jibei,
                record.department_jibei,
                record.homework_count,
                record.weekly_active_teachers,
                record.weekly_total_teachers,
                record.weekly_overall_activity,
                record.weekly_active_ratio,
                _date,
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.font = _CELL_FONT
                cell.alignment = _CELL_ALIGN
                cell.border = _THIN_BORDER

        col_widths = [18, 10, 12, 12, 12, 12, 12, 14, 14, 14, 12, 12]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    range_label = _safe_filename("-".join(week_labels[:1] + week_labels[-1:]))
    filename = f"weekly_{year}_{range_label}_{timestamp}.xlsx"
    filepath = _EXPORT_DIR / filename
    wb.save(str(filepath))

    return str(filepath)


# 月表表头定义
MONTHLY_HEADERS_ROW1 = [
    ("学校名称", 1), ("月次", 1),
    ("平台使用率", 4),
    ("集备", 4),
    ("组卷", 4),
    ("活跃度占比", 3),
    ("作业次数", 1),
    ("日期", 1),
]
MONTHLY_HEADERS_ROW2 = [
    "", "",  # 学校名称, 月次 (rowspan)
    "整体", "高中", "初中", "小学",  # 平台使用率
    "整体", "高中", "初中", "小学",  # 集备
    "整体", "高中", "初中", "小学",  # 组卷
    "日活", "周活", "月活",  # 活跃度占比
    "",  # 作业次数 (rowspan)
    "",  # 日期 (rowspan)
]


def export_monthly(
    records: list,
    year: int,
    month_number: str,
) -> str:
    """
    将月度记录导出为 Excel 文件（双行合并表头）。

    Returns:
        导出文件的路径
    """
    _EXPORT_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}年{month_number or '全部'}"[:31]

    total_cols = sum(cols for _, cols in MONTHLY_HEADERS_ROW1)

    # 标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"{year}年{month_number or '全部'}数据统计"
    title_cell.font = Font(name="微软雅黑", size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # 第一行表头（合并单元格）
    col = 1
    for header, span in MONTHLY_HEADERS_ROW1:
        if span > 1:
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + span - 1)
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER
        # Apply border to all merged cells
        for c in range(col, col + span):
            ws.cell(row=2, column=c).border = _THIN_BORDER
            ws.cell(row=2, column=c).font = _HEADER_FONT
            ws.cell(row=2, column=c).fill = _HEADER_FILL
            ws.cell(row=2, column=c).alignment = _HEADER_ALIGN
        col += span

    # 第二行表头（子列）
    col = 1
    for i, sub_header in enumerate(MONTHLY_HEADERS_ROW2):
        # Skip rowspan columns (学校名称, 月次, 作业次数)
        if i < 2:
            # Merge row 2-3 for rowspan columns, style already set in first pass
            ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
            col += 1
            continue
        if i == 17:  # 作业次数 (rowspan)
            top_cell = ws.cell(row=2, column=col)
            top_cell.font = _HEADER_FONT
            top_cell.fill = _HEADER_FILL
            top_cell.alignment = _HEADER_ALIGN
            top_cell.border = _THIN_BORDER
            ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
            col += 1
            continue
        if i == 18:  # 日期 (last column, rowspan)
            top_cell = ws.cell(row=2, column=col)
            top_cell.font = _HEADER_FONT
            top_cell.fill = _HEADER_FILL
            top_cell.alignment = _HEADER_ALIGN
            top_cell.border = _THIN_BORDER
            ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
            col += 1
            continue
        if sub_header:
            cell = ws.cell(row=3, column=col)
            cell.value = sub_header
            cell.font = Font(name="微软雅黑", size=9, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="5B8BD6", end_color="5B8BD6", fill_type="solid")
            cell.alignment = _HEADER_ALIGN
            cell.border = _THIN_BORDER
        col += 1

    # 写入数据
    for row_idx, record in enumerate(records, 4):
        values = [
            record.school_name,
            record.month_number,
            record.platform_usage,
            record.platform_usage_hs,
            record.platform_usage_ms,
            record.platform_usage_ps,
            record.overall_jibei,
            record.jibei_hs,
            record.jibei_ms,
            record.jibei_ps,
            record.zujuan,
            record.zujuan_hs,
            record.zujuan_ms,
            record.zujuan_ps,
            record.daily_active_ratio,
            record.weekly_active_ratio,
            record.monthly_active_ratio,
            record.homework_count,
            (record.collected_at.replace('T', ' ')[:16] if record.collected_at and 'T' in record.collected_at else (record.collected_at[:16] if record.collected_at else '')),
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = _CELL_FONT
            cell.alignment = _CELL_ALIGN
            cell.border = _THIN_BORDER

    # 调整列宽
    col_widths = [18, 8] + [10]*4 + [10]*4 + [10]*4 + [10]*3 + [10] + [12]
    for i, width in enumerate(col_widths, 1):
        if i <= total_cols:
            ws.column_dimensions[get_column_letter(i)].width = width

    # 保存文件
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_label = _safe_filename(month_number) if month_number else "all"
    filename = _build_export_name(year, month_number) + f"_{timestamp}.xlsx"
    filepath = _EXPORT_DIR / filename
    wb.save(str(filepath))

    return str(filepath)
